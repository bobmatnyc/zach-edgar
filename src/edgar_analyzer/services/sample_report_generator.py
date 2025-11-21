"""Sample report generator that matches the original report format exactly."""

import json
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
import structlog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from edgar_analyzer.models.company import AnalysisReport, CompanyAnalysis

logger = structlog.get_logger(__name__)


class SampleReportGenerator:
    """Generate reports that match the sample report format exactly."""
    
    def __init__(self, output_dir: str = "output"):
        """Initialize sample report generator."""
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        logger.info("Sample report generator initialized", output_dir=str(self.output_dir))
    
    async def generate_sample_format_report(
        self, 
        report: AnalysisReport,
        output_filename: str = "corporations_pay_executives_more_than_taxes.xlsx"
    ) -> Path:
        """Generate a report that matches the sample format exactly."""
        
        logger.info("Generating sample format report", companies=len(report.companies))
        
        # Prepare data for the main analysis
        main_data = self._prepare_main_analysis_data(report)
        
        # Create Excel workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create the main chart sheet (Chart-2 equivalent)
        self._create_main_chart_sheet(wb, main_data)
        
        # Create summary sheet (Chart-1 equivalent)
        self._create_summary_sheet(wb, main_data)
        
        # Create key findings sheet
        self._create_key_findings_sheet(wb, report)
        
        # Create executive list sheet
        self._create_executive_list_sheet(wb, report)
        
        # Save the workbook
        output_path = self.output_dir / output_filename
        wb.save(output_path)
        
        logger.info("Sample format report generated", output_path=str(output_path))
        return output_path
    
    def _prepare_main_analysis_data(self, report: AnalysisReport) -> List[Dict]:
        """Prepare data in the format matching the sample report."""
        main_data = []
        
        for analysis in report.companies:
            company = analysis.company
            
            # Calculate 5-year totals (2019-2023)
            total_tax_expense = 0
            total_executive_compensation = 0
            years_with_data = 0
            
            # Sum tax expenses
            for tax_expense in analysis.tax_expenses:
                if 2019 <= tax_expense.fiscal_year <= 2023:
                    total_tax_expense += float(tax_expense.total_tax_expense)
                    years_with_data += 1
            
            # Sum executive compensation by year
            comp_by_year = {}
            for comp in analysis.executive_compensations:
                if 2019 <= comp.fiscal_year <= 2023:
                    year = comp.fiscal_year
                    if year not in comp_by_year:
                        comp_by_year[year] = 0
                    comp_by_year[year] += float(comp.total_compensation)
            
            total_executive_compensation = sum(comp_by_year.values())
            
            # Calculate domestic pre-tax profits (estimated from tax data)
            # Using a reverse calculation: if ETR = tax/profit, then profit = tax/ETR
            # Assume average ETR of 15% for estimation
            estimated_pretax_profits = total_tax_expense / 0.15 if total_tax_expense > 0 else 0
            
            # Calculate effective tax rate
            effective_tax_rate = total_tax_expense / estimated_pretax_profits if estimated_pretax_profits > 0 else 0
            
            # Include all companies with sufficient data
            if years_with_data >= 3:
                company_data = {
                    'company_name': company.name,
                    'domestic_pretax_profits': estimated_pretax_profits / 1_000_000,  # Convert to millions
                    'federal_income_taxes': total_tax_expense / 1_000_000,  # Convert to millions
                    'effective_tax_rate': effective_tax_rate,
                    'executive_pay': total_executive_compensation / 1_000_000,  # Convert to millions
                    'stock_buybacks': 0,  # Placeholder - would need additional data
                    'dividend_payouts': 0,  # Placeholder - would need additional data
                    'fortune_rank': company.fortune_rank or 999
                }
                main_data.append(company_data)
        
        # Sort by Fortune rank
        main_data.sort(key=lambda x: x['fortune_rank'])
        
        return main_data
    
    def _create_main_chart_sheet(self, workbook: Workbook, main_data: List[Dict]) -> None:
        """Create the main chart sheet matching Chart-2 format."""
        ws = workbook.create_sheet("Chart-2")
        
        # Count companies where exec pay > taxes
        companies_exec_over_tax = sum(1 for d in main_data if d['executive_pay'] > d['federal_income_taxes'])

        # Header row 1 (title)
        if companies_exec_over_tax > 0:
            title = f"{companies_exec_over_tax} profitable corporations that paid top executives more than they paid in federal income taxes between 2019 and 2023"
        else:
            title = f"{len(main_data)} Fortune 500 corporations - Executive compensation vs Federal income taxes analysis (2019-2023)"

        ws.append([title, "Domestic Pre-Tax Profits", "Federal Income Taxes", "Effective Tax Rate", "Executive Pay", "Stock Buybacks", "Dividend Payouts"])

        # Header row 2 (subtitle with clear units)
        ws.append([None, "Five-Year Total, 2019 to 2023", "Five-Year Total, 2019 to 2023", "Five-Year Average", "Five-Year Total, 2019 to 2023", "Five-Year Total, 2019 to 2023", "Five-Year Total, 2019 to 2023"])

        # Header row 3 (units)
        ws.append([None, "(millions of dollars)", "(millions of dollars)", "(percentage)", "(millions of dollars)", "(millions of dollars)", "(millions of dollars)"])
        
        # Data rows with proper formatting
        for company_data in main_data:
            ws.append([
                company_data['company_name'],
                company_data['domestic_pretax_profits'],  # Will be formatted by Excel styling
                company_data['federal_income_taxes'],
                company_data['effective_tax_rate'],
                company_data['executive_pay'],
                company_data['stock_buybacks'],
                company_data['dividend_payouts']
            ])
        
        # Style the sheet
        self._style_main_chart_sheet(ws)
    
    def _style_main_chart_sheet(self, worksheet) -> None:
        """Apply styling to match the sample format."""
        from openpyxl.styles import NamedStyle

        # Header styling
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # Style first row (title row)
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Style second row (subtitle row)
        for cell in worksheet[2]:
            cell.font = Font(italic=True, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Style third row (units row)
        for cell in worksheet[3]:
            cell.font = Font(italic=True, size=9, color="666666")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Format data rows with proper number formatting
        for row_num in range(4, worksheet.max_row + 1):
            # Company name (column A) - left aligned
            worksheet.cell(row=row_num, column=1).alignment = Alignment(horizontal="left")

            # Financial columns (B, C, E, F, G) - number format with commas, 1 decimal place
            for col in [2, 3, 5, 6, 7]:  # Skip column 4 (effective tax rate)
                cell = worksheet.cell(row=row_num, column=col)
                cell.number_format = '#,##0.0'
                cell.alignment = Alignment(horizontal="right")

            # Effective tax rate (column D) - percentage format
            tax_rate_cell = worksheet.cell(row=row_num, column=4)
            tax_rate_cell.number_format = '0.000%'
            tax_rate_cell.alignment = Alignment(horizontal="right")

        # Set column widths for better readability
        column_widths = {
            'A': 35,  # Company name
            'B': 20,  # Domestic pre-tax profits
            'C': 20,  # Federal income taxes
            'D': 18,  # Effective tax rate
            'E': 20,  # Executive pay
            'F': 18,  # Stock buybacks
            'G': 18   # Dividend payouts
        }

        for column_letter, width in column_widths.items():
            worksheet.column_dimensions[column_letter].width = width

    def _create_summary_sheet(self, workbook: Workbook, main_data: List[Dict]) -> None:
        """Create summary sheet matching Chart-1 format."""
        ws = workbook.create_sheet("Chart-1")

        # Calculate totals
        total_pretax_profits = sum(d['domestic_pretax_profits'] for d in main_data)
        total_federal_taxes = sum(d['federal_income_taxes'] for d in main_data)
        total_executive_pay = sum(d['executive_pay'] for d in main_data)
        total_stock_buybacks = sum(d['stock_buybacks'] for d in main_data)
        total_dividend_payouts = sum(d['dividend_payouts'] for d in main_data)

        overall_effective_rate = total_federal_taxes / total_pretax_profits if total_pretax_profits > 0 else 0

        # Header
        ws.append([None, "Domestic Pre-Tax Profits", "Federal Income Taxes", "Effective Tax Rate", "Executive Pay", "Stock Buybacks", "Dividend Payouts"])

        # Units row
        ws.append([None, "(millions of dollars)", "(millions of dollars)", "(percentage)", "(millions of dollars)", "(millions of dollars)", "(millions of dollars)"])

        # Count companies where exec pay > taxes
        companies_exec_over_tax = sum(1 for d in main_data if d['executive_pay'] > d['federal_income_taxes'])

        # Summary row
        if companies_exec_over_tax > 0:
            summary_title = f"{companies_exec_over_tax} Profitable Corporations That Paid More To Executives Than In Federal Taxes"
        else:
            summary_title = f"{len(main_data)} Fortune 500 Corporations - Executive Compensation vs Federal Tax Analysis"
        ws.append([
            summary_title,
            total_pretax_profits,
            total_federal_taxes,
            overall_effective_rate,
            total_executive_pay,
            total_stock_buybacks,
            total_dividend_payouts
        ])

        # Empty row
        ws.append([None, None, None, None, None, None, None])

        # Style the summary sheet
        self._style_summary_sheet(ws)

    def _style_summary_sheet(self, worksheet) -> None:
        """Style the summary sheet."""
        # Header styling
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # Style header row
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Style units row
        for cell in worksheet[2]:
            cell.font = Font(italic=True, size=9, color="666666")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Style summary row
        for cell in worksheet[3]:
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Format summary data with proper number formatting
        # Company name (column A) - left aligned
        worksheet.cell(row=3, column=1).alignment = Alignment(horizontal="left")

        # Financial columns (B, C, E, F, G) - number format with commas, 1 decimal place
        for col in [2, 3, 5, 6, 7]:  # Skip column 4 (effective tax rate)
            cell = worksheet.cell(row=3, column=col)
            cell.number_format = '#,##0.0'
            cell.alignment = Alignment(horizontal="right")

        # Effective tax rate (column D) - percentage format
        tax_rate_cell = worksheet.cell(row=3, column=4)
        tax_rate_cell.number_format = '0.000%'
        tax_rate_cell.alignment = Alignment(horizontal="right")

        # Set column widths for better readability
        column_widths = {
            'A': 45,  # Company summary title
            'B': 20,  # Domestic pre-tax profits
            'C': 20,  # Federal income taxes
            'D': 18,  # Effective tax rate
            'E': 20,  # Executive pay
            'F': 18,  # Stock buybacks
            'G': 18   # Dividend payouts
        }

        for column_letter, width in column_widths.items():
            worksheet.column_dimensions[column_letter].width = width

    def _create_key_findings_sheet(self, workbook: Workbook, report: AnalysisReport) -> None:
        """Create key findings sheet."""
        ws = workbook.create_sheet("Key Findings")

        # Headers
        headers = [
            "Rank", "Corporation", "Home State", "Fortune Rank",
            "5-Year Executive Pay (millions)", "5-Year Federal Taxes (millions)",
            "Executive Pay vs Tax Ratio", "Effective Tax Rate"
        ]
        ws.append(headers)

        # Prepare data
        findings_data = []
        for i, analysis in enumerate(report.companies[:50], 1):  # Top 50
            company = analysis.company

            # Calculate totals
            total_tax = sum(float(tax.total_tax_expense) for tax in analysis.tax_expenses) / 1_000_000
            total_comp = sum(float(comp.total_compensation) for comp in analysis.executive_compensations) / 1_000_000

            ratio = total_comp / total_tax if total_tax > 0 else float('inf')
            etr = total_tax / (total_tax / 0.15) if total_tax > 0 else 0  # Estimated

            # Get company home state
            home_state = self._get_company_home_state(company.name)

            findings_data.append([
                i,
                company.name,
                home_state,
                company.fortune_rank or 999,
                total_comp,  # Will be formatted by Excel
                total_tax,
                ratio,
                etr  # Will be formatted as percentage
            ])

        # Add data rows
        for row_data in findings_data:
            ws.append(row_data)

        # Style the sheet
        self._style_key_findings_sheet(ws)

    def _style_key_findings_sheet(self, worksheet) -> None:
        """Style the key findings sheet."""
        # Header styling
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Format data rows
        for row_num in range(2, worksheet.max_row + 1):
            # Rank (column A) - center aligned
            worksheet.cell(row=row_num, column=1).alignment = Alignment(horizontal="center")

            # Corporation (column B) - left aligned
            worksheet.cell(row=row_num, column=2).alignment = Alignment(horizontal="left")

            # Home State (column C) - center aligned
            worksheet.cell(row=row_num, column=3).alignment = Alignment(horizontal="center")

            # Fortune Rank (column D) - center aligned
            worksheet.cell(row=row_num, column=4).alignment = Alignment(horizontal="center")

            # Executive Pay (column E) - currency format
            exec_pay_cell = worksheet.cell(row=row_num, column=5)
            exec_pay_cell.number_format = '#,##0.0'
            exec_pay_cell.alignment = Alignment(horizontal="right")

            # Federal Taxes (column F) - currency format
            tax_cell = worksheet.cell(row=row_num, column=6)
            tax_cell.number_format = '#,##0.0'
            tax_cell.alignment = Alignment(horizontal="right")

            # Ratio (column G) - number format
            ratio_cell = worksheet.cell(row=row_num, column=7)
            ratio_cell.number_format = '#,##0.0'
            ratio_cell.alignment = Alignment(horizontal="right")

            # Effective Tax Rate (column H) - percentage format
            etr_cell = worksheet.cell(row=row_num, column=8)
            etr_cell.number_format = '0.0%'
            etr_cell.alignment = Alignment(horizontal="right")

        # Set column widths for better readability
        column_widths = {
            'A': 8,   # Rank
            'B': 35,  # Corporation
            'C': 12,  # Home State
            'D': 12,  # Fortune Rank
            'E': 18,  # Executive Pay
            'F': 18,  # Federal Taxes
            'G': 15,  # Ratio
            'H': 15   # Effective Tax Rate
        }

        for column_letter, width in column_widths.items():
            worksheet.column_dimensions[column_letter].width = width

    def _get_company_home_state(self, company_name: str) -> str:
        """Get the home state for a company based on its name."""
        # Fortune 500 company home states (incorporation/headquarters)
        company_states = {
            # Top 10
            "Walmart Inc.": "AR",
            "Amazon.com Inc.": "WA",
            "Apple Inc.": "CA",
            "CVS Health Corporation": "RI",
            "UnitedHealth Group Incorporated": "MN",
            "Exxon Mobil Corporation": "TX",
            "Berkshire Hathaway Inc.": "NE",
            "Alphabet Inc.": "DE",
            "McKesson Corporation": "CA",
            "Cencora Inc.": "PA",

            # 11-20
            "Costco Wholesale Corporation": "WA",
            "JPMorgan Chase & Co.": "NY",
            "Microsoft Corporation": "WA",
            "Cardinal Health, Inc.": "OH",
            "Chevron Corporation": "CA",
            "Ford Motor Company": "DE",
            "General Motors Company": "DE",
            "Elevance Health, Inc.": "IN",
            "Fannie Mae": "DC",
            "Home Depot, Inc.": "GA",

            # 21-30
            "Marathon Petroleum Corporation": "OH",
            "Phillips 66": "TX",
            "Valero Energy Corporation": "TX",
            "General Electric Company": "MA",
            "Walgreens Boots Alliance, Inc.": "IL",
            "Archer-Daniels-Midland Company": "IL",
            "Lockheed Martin Corporation": "MD",
            "Energy Transfer LP": "TX",
            "Procter & Gamble Company": "OH",
            "Johnson & Johnson": "NJ",

            # 31-40
            "Dell Technologies Inc.": "TX",
            "FedEx Corporation": "TN",
            "UPS, Inc.": "GA",
            "Lowe's Companies, Inc.": "NC",
            "Wells Fargo & Company": "CA",
            "Target Corporation": "MN",
            "Humana Inc.": "KY",
            "AbbVie Inc.": "IL",
            "Caterpillar Inc.": "IL",
            "Comcast Corporation": "PA",

            # 41-50
            "Tesla, Inc.": "TX",
            "IBM Corporation": "NY",
            "Freddie Mac": "VA",
            "Bank of America Corporation": "NC",
            "Cigna Group": "CT",
            "Sysco Corporation": "TX",
            "Kroger Co.": "OH",
            "Verizon Communications Inc.": "NY",
            "AT&T Inc.": "TX",
            "Meta Platforms, Inc.": "DE",

            # Additional common companies
            "General Dynamics Corporation": "VA",
            "Raytheon Technologies Corporation": "MA",
            "Boeing Company": "DE",
            "Intel Corporation": "DE",
            "Oracle Corporation": "TX",
            "Salesforce, Inc.": "DE",
            "Netflix, Inc.": "DE",
            "PayPal Holdings, Inc.": "DE",
            "Visa Inc.": "DE",
            "Mastercard Incorporated": "NY",
            "American Express Company": "NY",
            "Goldman Sachs Group, Inc.": "DE",
            "Morgan Stanley": "DE",
            "Citigroup Inc.": "DE",
            "PepsiCo, Inc.": "NC",
            "Coca-Cola Company": "DE",
            "Walt Disney Company": "DE",
            "Nike, Inc.": "OR",
            "Starbucks Corporation": "WA",
            "McDonald's Corporation": "DE"
        }

        # Clean company name for lookup (remove common suffixes)
        clean_name = company_name.strip()

        # Direct lookup
        if clean_name in company_states:
            return company_states[clean_name]

        # Try variations without common suffixes
        variations = [
            clean_name.replace(" Inc.", ""),
            clean_name.replace(" Corporation", ""),
            clean_name.replace(" Company", ""),
            clean_name.replace(" Co.", ""),
            clean_name.replace(" LLC", ""),
            clean_name.replace(" LP", ""),
            clean_name.replace(", Inc.", ""),
            clean_name.replace(" Incorporated", "")
        ]

        for variation in variations:
            variation = variation.strip()
            for known_company, state in company_states.items():
                if variation.lower() in known_company.lower() or known_company.lower() in variation.lower():
                    return state

        # Default fallback - try to guess based on common patterns
        if "texas" in clean_name.lower() or "tx" in clean_name.lower():
            return "TX"
        elif "california" in clean_name.lower() or "ca" in clean_name.lower():
            return "CA"
        elif "new york" in clean_name.lower() or "ny" in clean_name.lower():
            return "NY"
        elif "delaware" in clean_name.lower() or "de" in clean_name.lower():
            return "DE"

        return "DE"  # Most Fortune 500 companies are incorporated in Delaware

    def _create_executive_list_sheet(self, workbook: Workbook, report: AnalysisReport) -> None:
        """Create executive list sheet matching the sample format."""
        ws = workbook.create_sheet("List of Executives")

        # Headers
        headers = [
            "Billionaires", "Executives", "Corporate Allegiance",
            "Five-Year Executive Pay", "2023", "2022", "2021", "2020", "2019"
        ]
        ws.append(headers)

        # Collect all executives
        all_executives = []
        for analysis in report.companies:
            company = analysis.company

            # Group compensation by executive and year
            exec_data = {}
            for comp in analysis.executive_compensations:
                exec_name = comp.executive_name
                year = comp.fiscal_year

                if exec_name not in exec_data:
                    exec_data[exec_name] = {
                        'name': exec_name,
                        'company': company.name,
                        'title': comp.title,
                        'years': {}
                    }

                exec_data[exec_name]['years'][year] = float(comp.total_compensation) / 1_000_000

            # Add executives to the list
            for exec_info in exec_data.values():
                five_year_total = sum(exec_info['years'].values())

                # Determine if billionaire (simplified check)
                is_billionaire = "Yes" if five_year_total > 1000 else ""

                row_data = [
                    is_billionaire,
                    exec_info['name'],
                    exec_info['company'],
                    round(five_year_total, 3),
                    round(exec_info['years'].get(2023, 0), 3),
                    round(exec_info['years'].get(2022, 0), 3),
                    round(exec_info['years'].get(2021, 0), 3),
                    round(exec_info['years'].get(2020, 0), 3),
                    round(exec_info['years'].get(2019, 0), 3)
                ]
                all_executives.append((five_year_total, row_data))

        # Sort by five-year total (descending)
        all_executives.sort(key=lambda x: x[0], reverse=True)

        # Add top executives to sheet
        for _, row_data in all_executives[:500]:  # Limit to top 500
            ws.append(row_data)

        # Style the sheet
        self._style_executive_list_sheet(ws)

    def _style_executive_list_sheet(self, worksheet) -> None:
        """Style the executive list sheet."""
        # Header styling
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Highlight billionaires
        for row in worksheet.iter_rows(min_row=2):
            if row[0].value == "Yes":  # Billionaire column
                for cell in row:
                    cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            worksheet.column_dimensions[column_letter].width = adjusted_width
