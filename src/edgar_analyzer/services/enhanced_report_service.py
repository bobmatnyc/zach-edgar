"""Enhanced report generation service matching sample Excel format."""

import json
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
import structlog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from edgar_analyzer.config.settings import ConfigService
from edgar_analyzer.models.company import AnalysisReport, CompanyAnalysis
from edgar_analyzer.services.historical_analysis_service import HistoricalAnalysisService
from edgar_analyzer.services.interfaces import IDataExtractionService, IReportService
from edgar_analyzer.validation.quality_reporter import QualityReporter

logger = structlog.get_logger(__name__)


class EnhancedReportService(IReportService):
    """Enhanced report generation service matching sample format."""

    def __init__(
        self,
        data_extraction_service: IDataExtractionService,
        historical_analysis_service: HistoricalAnalysisService,
        config: ConfigService
    ):
        """Initialize enhanced report service."""
        self._data_extraction = data_extraction_service
        self._historical_analysis = historical_analysis_service
        self._config = config
        self._output_dir = Path(config.settings.output_dir)
        self._output_dir.mkdir(parents=True, exist_ok=True)

        # Initialize quality reporter
        self._quality_reporter = QualityReporter(str(self._output_dir))

        logger.info("Enhanced report service initialized", output_dir=str(self._output_dir))

    async def generate_analysis_report(
        self, companies: List[str], year: int
    ) -> AnalysisReport:
        """Generate comprehensive analysis report for multiple companies."""
        logger.info("Starting enhanced analysis report generation", companies=len(companies), year=year)

        # Generate 5-year historical analysis
        years = [year - 4, year - 3, year - 2, year - 1, year]  # 5-year period
        report = AnalysisReport(target_year=year)

        for cik in companies:
            try:
                logger.info("Analyzing company with historical data", cik=cik)

                # Extract multi-year analysis
                company_analysis = await self._historical_analysis.extract_multi_year_analysis(cik, years)

                if company_analysis:
                    report.add_company_analysis(company_analysis)
                    logger.info(
                        "Company analysis added to report",
                        cik=cik,
                        company=company_analysis.company.name
                    )
                else:
                    logger.warning("Failed to analyze company", cik=cik)

            except Exception as e:
                logger.error("Error analyzing company", cik=cik, error=str(e))
                continue

        logger.info(
            "Enhanced analysis report generation completed",
            total_companies=report.total_companies,
            year=year
        )

        return report

    async def export_to_excel(self, report: AnalysisReport, filepath: str) -> None:
        """Export report to Excel file matching sample format."""
        try:
            # Create Excel workbook
            wb = Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # Create sheets matching sample format
            await self._create_main_analysis_sheet(wb, report)
            await self._create_summary_sheet(wb, report)
            await self._create_trends_sheet(wb, report)
            await self._create_methodology_sheet(wb, report)

            # Save workbook
            output_path = self._output_dir / filepath
            wb.save(output_path)

            logger.info("Enhanced Excel report exported", filepath=str(output_path))

        except Exception as e:
            logger.error("Failed to export enhanced Excel report", filepath=filepath, error=str(e))
            raise

    async def export_to_json(self, report: AnalysisReport, filepath: str) -> None:
        """Export report to JSON file."""
        try:
            output_path = self._output_dir / filepath

            # Convert report to dict with enhanced data
            report_data = await self._create_enhanced_json_data(report)

            # Write to JSON file
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(report_data, f, indent=2, default=str)

            logger.info("Enhanced JSON report exported", filepath=str(output_path))

        except Exception as e:
            logger.error("Failed to export enhanced JSON report", filepath=filepath, error=str(e))
            raise

    async def generate_quality_report(
        self,
        report: AnalysisReport,
        include_in_main_report: bool = True
    ) -> Dict[str, any]:
        """Generate comprehensive data quality report."""

        logger.info("Generating data quality report", companies=len(report.companies))

        # Generate quality report
        quality_data = await self._quality_reporter.generate_quality_report(
            report.companies,
            f"quality_report_{report.target_year}"
        )

        # Add quality summary to main report if requested
        if include_in_main_report:
            report.quality_summary = {
                "overall_score": quality_data["summary"]["overall_quality_score"],
                "overall_grade": quality_data["summary"]["overall_grade"],
                "critical_issues": quality_data["summary"]["critical_issues"],
                "total_validations": quality_data["summary"]["total_validations"],
                "companies_by_grade": quality_data["summary"]["companies_by_grade"]
            }

        logger.info(
            "Data quality report generated",
            overall_score=quality_data["summary"]["overall_quality_score"],
            overall_grade=quality_data["summary"]["overall_grade"]
        )

        return quality_data

    async def export_with_quality_validation(
        self,
        report: AnalysisReport,
        output_filename: str
    ) -> Tuple[Path, Dict[str, any]]:
        """Export report with integrated quality validation."""

        # Generate quality report
        quality_data = await self.generate_quality_report(report, include_in_main_report=True)

        # Export main report
        if output_filename.endswith('.xlsx'):
            report_path = await self.export_to_excel(report, output_filename)
        elif output_filename.endswith('.json'):
            report_path = await self.export_to_json(report, output_filename)
        else:
            raise ValueError("Output file must be .xlsx or .json")

        logger.info(
            "Report exported with quality validation",
            report_path=str(report_path),
            quality_score=quality_data["summary"]["overall_quality_score"]
        )

        return report_path, quality_data

    async def _create_enhanced_dataframe(self, report: AnalysisReport) -> pd.DataFrame:
        """Create enhanced DataFrame matching sample format."""
        data = []
        years = [report.target_year - 4, report.target_year - 3, report.target_year - 2,
                report.target_year - 1, report.target_year]

        for company_analysis in report.companies:
            company = company_analysis.company

            # Get tax expenses by year
            tax_by_year = {}
            for tax_expense in company_analysis.tax_expenses:
                tax_by_year[tax_expense.fiscal_year] = float(tax_expense.total_tax_expense)

            # Get executive compensation by year
            comp_by_year = company_analysis.total_executive_compensation

            # Calculate 5-year totals
            total_tax_5yr = sum(tax_by_year.get(year, 0) for year in years)
            total_comp_5yr = sum(float(comp_by_year.get(year, 0)) for year in years)

            # Calculate ratios
            current_year_tax = tax_by_year.get(report.target_year, 0)
            current_year_comp = float(comp_by_year.get(report.target_year, 0))

            ratio_current = current_year_comp / current_year_tax if current_year_tax > 0 else 0
            ratio_5yr = total_comp_5yr / total_tax_5yr if total_tax_5yr > 0 else 0

            row = {
                'Rank': company.fortune_rank or 999,
                'Company': company.name,
                'Ticker': company.ticker or 'N/A',
                'Sector': company.sector or 'Unknown',
                'Industry': company.industry or 'Unknown',

                # 5-year data
                f'Tax Expense {years[0]}': tax_by_year.get(years[0], 0),
                f'Tax Expense {years[1]}': tax_by_year.get(years[1], 0),
                f'Tax Expense {years[2]}': tax_by_year.get(years[2], 0),
                f'Tax Expense {years[3]}': tax_by_year.get(years[3], 0),
                f'Tax Expense {years[4]}': tax_by_year.get(years[4], 0),

                f'Exec Comp {years[0]}': float(comp_by_year.get(years[0], 0)),
                f'Exec Comp {years[1]}': float(comp_by_year.get(years[1], 0)),
                f'Exec Comp {years[2]}': float(comp_by_year.get(years[2], 0)),
                f'Exec Comp {years[3]}': float(comp_by_year.get(years[3], 0)),
                f'Exec Comp {years[4]}': float(comp_by_year.get(years[4], 0)),

                # Totals and ratios
                'Total Tax Expense (5yr)': total_tax_5yr,
                'Total Exec Comp (5yr)': total_comp_5yr,
                f'Comp/Tax Ratio {report.target_year}': ratio_current,
                'Comp/Tax Ratio (5yr avg)': ratio_5yr,
                'Comp Exceeds Tax': 'Yes' if ratio_current > 1 else 'No',
            }

            data.append(row)

        df = pd.DataFrame(data)

        # Sort by Fortune ranking
        if not df.empty:
            df = df.sort_values('Rank', ascending=True)

        return df

    async def _create_main_analysis_sheet(self, workbook: Workbook, report: AnalysisReport) -> None:
        """Create main analysis sheet matching sample format."""
        ws = workbook.create_sheet("Executive Compensation vs Tax Analysis")

        # Create comprehensive DataFrame
        df = await self._create_enhanced_dataframe(report)

        # Add title row
        title = f"Executive Compensation vs Tax Expense Analysis - Fortune 500 Companies ({report.target_year})"
        ws.append([title])
        ws.merge_cells('A1:T1')

        # Style title
        title_cell = ws['A1']
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add empty row
        ws.append([])

        # Add headers and data
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Style header row
        header_row = 3
        for cell in ws[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # Auto-adjust column widths and add borders
        for col_idx, column in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = chr(64 + col_idx)  # A, B, C, etc.

            for cell in column:
                try:
                    # Skip merged cells
                    if hasattr(cell, 'coordinate') and cell.coordinate in ws.merged_cells:
                        continue

                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))

                    # Add borders to data cells
                    if cell.row > 2:
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )

                        # Format numbers
                        if isinstance(cell.value, (int, float)) and cell.value > 1000:
                            cell.number_format = '#,##0'

                except:
                    pass

            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width

    async def _create_summary_sheet(self, workbook: Workbook, report: AnalysisReport) -> None:
        """Create summary sheet."""
        ws = workbook.create_sheet("Summary")

        # Add summary statistics
        stats = report.summary_statistics

        summary_data = [
            ["Enhanced Analysis Summary", ""],
            ["Report Date", stats["report_date"].strftime("%Y-%m-%d %H:%M:%S")],
            ["Target Year", stats["target_year"]],
            ["Analysis Period", f"{stats['target_year']-4}-{stats['target_year']} (5 years)"],
            ["Total Companies Analyzed", stats["total_companies"]],
            ["Companies with Higher Compensation", stats["companies_with_higher_compensation"]],
            ["Percentage with Higher Compensation", f"{stats['percentage_higher_compensation']:.1f}%"],
        ]

        for row in summary_data:
            ws.append(row)

        # Style summary sheet
        header_font = Font(bold=True, size=14)
        ws['A1'].font = header_font

    async def _create_trends_sheet(self, workbook: Workbook, report: AnalysisReport) -> None:
        """Create trends analysis sheet."""
        ws = workbook.create_sheet("Trends")

        ws.append(["Historical Trends Analysis"])
        ws.append(["This sheet would contain trend analysis charts and data"])
        ws.append(["Implementation pending based on requirements"])

    async def _create_methodology_sheet(self, workbook: Workbook, report: AnalysisReport) -> None:
        """Create methodology sheet."""
        ws = workbook.create_sheet("Methodology")

        methodology_data = [
            ["Analysis Methodology"],
            [""],
            ["Data Sources:"],
            ["• SEC EDGAR API for financial data"],
            ["• Form 10-K for tax expense information"],
            ["• DEF 14A proxy statements for executive compensation"],
            [""],
            ["Analysis Period:"],
            [f"• 5-year historical analysis ({report.target_year-4}-{report.target_year})"],
            ["• Annual data extraction and comparison"],
            [""],
            ["Metrics Calculated:"],
            ["• Total executive compensation by year"],
            ["• Income tax expense by year"],
            ["• Compensation to tax expense ratios"],
            ["• 5-year averages and trends"],
        ]

        for row in methodology_data:
            ws.append(row)

        # Style methodology sheet
        header_font = Font(bold=True, size=14)
        ws['A1'].font = header_font

    async def _create_enhanced_json_data(self, report: AnalysisReport) -> Dict:
        """Create enhanced JSON data structure."""
        return {
            "report_metadata": {
                "report_date": report.report_date.isoformat(),
                "target_year": report.target_year,
                "analysis_period": f"{report.target_year-4}-{report.target_year}",
                "total_companies": report.total_companies,
                "report_type": "Enhanced Fortune 500 Analysis"
            },
            "summary_statistics": report.summary_statistics,
            "companies": [
                {
                    "company_info": analysis.company.dict(),
                    "tax_expenses": [tax.dict() for tax in analysis.tax_expenses],
                    "executive_compensations": [comp.dict() for comp in analysis.executive_compensations],
                    "analysis_metrics": {
                        "total_compensation_by_year": {
                            str(year): float(amount)
                            for year, amount in analysis.total_executive_compensation.items()
                        },
                        "compensation_vs_tax_ratios": {
                            str(year): float(ratio) if ratio else None
                            for year, ratio in analysis.compensation_vs_tax_ratio.items()
                        }
                    }
                }
                for analysis in report.companies
            ]
        }