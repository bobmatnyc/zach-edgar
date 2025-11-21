"""Data quality reporting and validation summary service."""

import json
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
import structlog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from edgar_analyzer.models.company import CompanyAnalysis
from edgar_analyzer.validation.data_validator import DataValidator, ValidationResult
from edgar_analyzer.validation.sanity_checker import SanityChecker
from edgar_analyzer.validation.source_verifier import SourceVerifier

logger = structlog.get_logger(__name__)


class QualityReporter:
    """Generate comprehensive data quality reports."""
    
    def __init__(self, output_dir: str = "output"):
        """Initialize quality reporter."""
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Initialize validation components
        self.data_validator = DataValidator()
        self.sanity_checker = SanityChecker()
        self.source_verifier = SourceVerifier()
        
        logger.info("Quality reporter initialized", output_dir=str(self.output_dir))
    
    async def generate_quality_report(
        self, 
        analyses: List[CompanyAnalysis],
        report_name: str = "data_quality_report"
    ) -> Dict[str, any]:
        """Generate comprehensive data quality report."""
        
        logger.info("Generating data quality report", companies=len(analyses))
        
        all_validations = []
        company_scores = {}
        
        # Validate each company's data
        for analysis in analyses:
            company_validations = await self._validate_company_analysis(analysis)
            all_validations.extend(company_validations)
            
            # Calculate company quality score
            company_score = self._calculate_quality_score(company_validations)
            company_scores[analysis.company.name] = company_score
        
        # Generate summary statistics
        summary = self._generate_summary_statistics(all_validations, company_scores)
        
        # Create detailed report
        report_data = {
            "metadata": {
                "generated_at": datetime.now().isoformat(),
                "total_companies": len(analyses),
                "total_validations": len(all_validations),
                "report_name": report_name
            },
            "summary": summary,
            "company_scores": company_scores,
            "detailed_validations": [
                {
                    "company": validation.get("company_name"),
                    "validation": validation["result"].dict()
                }
                for validation in all_validations
            ]
        }
        
        # Export reports
        await self._export_quality_reports(report_data, report_name)
        
        logger.info(
            "Data quality report generated",
            overall_score=summary["overall_quality_score"],
            critical_issues=summary["critical_issues"],
            warnings=summary["warnings"]
        )
        
        return report_data
    
    async def _validate_company_analysis(self, analysis: CompanyAnalysis) -> List[Dict]:
        """Validate a single company's analysis."""
        validations = []
        company_name = analysis.company.name
        
        # Validate tax expenses
        for tax_expense in analysis.tax_expenses:
            tax_validations = self.data_validator.validate_tax_expense(tax_expense)
            for validation in tax_validations:
                validations.append({
                    "company_name": company_name,
                    "data_type": "tax_expense",
                    "year": tax_expense.fiscal_year,
                    "result": validation
                })
        
        # Validate executive compensations
        for compensation in analysis.executive_compensations:
            comp_validations = self.data_validator.validate_executive_compensation(compensation)
            for validation in comp_validations:
                validations.append({
                    "company_name": company_name,
                    "data_type": "executive_compensation",
                    "year": compensation.fiscal_year,
                    "result": validation
                })
        
        # Perform sanity checks
        sanity_validations = self.sanity_checker.perform_comprehensive_sanity_check(analysis)
        for validation in sanity_validations:
            validations.append({
                "company_name": company_name,
                "data_type": "sanity_check",
                "year": None,
                "result": validation
            })
        
        # Perform source verification
        if analysis.tax_expenses and analysis.executive_compensations:
            # Use most recent year's data for verification
            latest_tax = max(analysis.tax_expenses, key=lambda x: x.fiscal_year)
            latest_comps = [c for c in analysis.executive_compensations 
                          if c.fiscal_year == latest_tax.fiscal_year]
            
            source_validations = await self.source_verifier.verify_against_benchmarks(
                analysis.company, latest_tax, latest_comps
            )
            for validation in source_validations:
                validations.append({
                    "company_name": company_name,
                    "data_type": "source_verification",
                    "year": latest_tax.fiscal_year,
                    "result": validation
                })
        
        # Perform spot checks
        spot_check_validations = await self.source_verifier.perform_spot_checks(
            analysis.company, 2023  # Default to 2023
        )
        for validation in spot_check_validations:
            validations.append({
                "company_name": company_name,
                "data_type": "spot_check",
                "year": None,
                "result": validation
            })
        
        return validations
    
    def _calculate_quality_score(self, validations: List[Dict]) -> Dict[str, any]:
        """Calculate quality score for a company."""
        if not validations:
            return {"score": 0.0, "grade": "F", "issues": 0}
        
        total_score = 0.0
        total_weight = 0.0
        critical_issues = 0
        errors = 0
        warnings = 0
        
        for validation in validations:
            result = validation["result"]
            confidence = result.confidence_score
            
            # Weight by severity
            if result.severity == "CRITICAL":
                weight = 1.0
                critical_issues += 1
            elif result.severity == "ERROR":
                weight = 0.8
                errors += 1
            elif result.severity == "WARNING":
                weight = 0.6
                warnings += 1
            else:  # INFO
                weight = 0.4
            
            total_score += confidence * weight
            total_weight += weight
        
        # Calculate weighted average
        if total_weight > 0:
            avg_score = total_score / total_weight
        else:
            avg_score = 1.0
        
        # Determine grade
        if avg_score >= 0.9:
            grade = "A"
        elif avg_score >= 0.8:
            grade = "B"
        elif avg_score >= 0.7:
            grade = "C"
        elif avg_score >= 0.6:
            grade = "D"
        else:
            grade = "F"
        
        return {
            "score": avg_score,
            "grade": grade,
            "critical_issues": critical_issues,
            "errors": errors,
            "warnings": warnings,
            "total_validations": len(validations)
        }

    def _generate_summary_statistics(
        self,
        all_validations: List[Dict],
        company_scores: Dict[str, Dict]
    ) -> Dict[str, any]:
        """Generate summary statistics for the quality report."""

        if not all_validations:
            return {
                "overall_quality_score": 0.0,
                "overall_grade": "F",
                "total_validations": 0,
                "critical_issues": 0,
                "errors": 0,
                "warnings": 0,
                "info_messages": 0,
                "companies_by_grade": {},
                "validation_types": {}
            }

        # Count by severity
        severity_counts = {"CRITICAL": 0, "ERROR": 0, "WARNING": 0, "INFO": 0}
        validation_types = {}

        for validation in all_validations:
            result = validation["result"]
            severity_counts[result.severity] += 1

            data_type = validation["data_type"]
            if data_type not in validation_types:
                validation_types[data_type] = {"total": 0, "passed": 0, "failed": 0}

            validation_types[data_type]["total"] += 1
            if result.is_valid:
                validation_types[data_type]["passed"] += 1
            else:
                validation_types[data_type]["failed"] += 1

        # Calculate overall quality score
        if company_scores:
            overall_score = sum(score["score"] for score in company_scores.values()) / len(company_scores)
        else:
            overall_score = 0.0

        # Determine overall grade
        if overall_score >= 0.9:
            overall_grade = "A"
        elif overall_score >= 0.8:
            overall_grade = "B"
        elif overall_score >= 0.7:
            overall_grade = "C"
        elif overall_score >= 0.6:
            overall_grade = "D"
        else:
            overall_grade = "F"

        # Count companies by grade
        companies_by_grade = {"A": 0, "B": 0, "C": 0, "D": 0, "F": 0}
        for score_data in company_scores.values():
            companies_by_grade[score_data["grade"]] += 1

        return {
            "overall_quality_score": overall_score,
            "overall_grade": overall_grade,
            "total_validations": len(all_validations),
            "critical_issues": severity_counts["CRITICAL"],
            "errors": severity_counts["ERROR"],
            "warnings": severity_counts["WARNING"],
            "info_messages": severity_counts["INFO"],
            "companies_by_grade": companies_by_grade,
            "validation_types": validation_types
        }

    async def _export_quality_reports(self, report_data: Dict, report_name: str) -> None:
        """Export quality reports in multiple formats."""

        # Export JSON report
        json_path = self.output_dir / f"{report_name}.json"
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, indent=2, default=str)

        # Export Excel report
        excel_path = self.output_dir / f"{report_name}.xlsx"
        await self._create_excel_quality_report(report_data, excel_path)

        logger.info(
            "Quality reports exported",
            json_path=str(json_path),
            excel_path=str(excel_path)
        )

    async def _create_excel_quality_report(self, report_data: Dict, excel_path: Path) -> None:
        """Create Excel quality report with multiple sheets."""

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create summary sheet
        self._create_summary_sheet(wb, report_data)

        # Create company scores sheet
        self._create_company_scores_sheet(wb, report_data)

        # Create detailed validations sheet
        self._create_validations_sheet(wb, report_data)

        # Save workbook
        wb.save(excel_path)

    def _create_summary_sheet(self, workbook: Workbook, report_data: Dict) -> None:
        """Create summary sheet in Excel report."""
        ws = workbook.create_sheet("Summary")

        summary = report_data["summary"]
        metadata = report_data["metadata"]

        # Title
        ws.append(["Data Quality Report Summary"])
        ws.merge_cells('A1:B1')
        ws['A1'].font = Font(bold=True, size=16)

        # Metadata
        ws.append([])
        ws.append(["Report Generated", metadata["generated_at"]])
        ws.append(["Total Companies", metadata["total_companies"]])
        ws.append(["Total Validations", summary["total_validations"]])
        ws.append([])

        # Overall scores
        ws.append(["Overall Quality Score", f"{summary['overall_quality_score']:.1%}"])
        ws.append(["Overall Grade", summary["overall_grade"]])
        ws.append([])

        # Issue counts
        ws.append(["Critical Issues", summary["critical_issues"]])
        ws.append(["Errors", summary["errors"]])
        ws.append(["Warnings", summary["warnings"]])
        ws.append(["Info Messages", summary["info_messages"]])
        ws.append([])

        # Companies by grade
        ws.append(["Companies by Grade", ""])
        for grade, count in summary["companies_by_grade"].items():
            ws.append([f"Grade {grade}", count])

    def _create_company_scores_sheet(self, workbook: Workbook, report_data: Dict) -> None:
        """Create company scores sheet."""
        ws = workbook.create_sheet("Company Scores")

        # Headers
        headers = ["Company", "Quality Score", "Grade", "Critical Issues", "Errors", "Warnings", "Total Validations"]
        ws.append(headers)

        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Data
        for company, score_data in report_data["company_scores"].items():
            ws.append([
                company,
                f"{score_data['score']:.1%}",
                score_data["grade"],
                score_data["critical_issues"],
                score_data["errors"],
                score_data["warnings"],
                score_data["total_validations"]
            ])

    def _create_validations_sheet(self, workbook: Workbook, report_data: Dict) -> None:
        """Create detailed validations sheet."""
        ws = workbook.create_sheet("Detailed Validations")

        # Headers
        headers = ["Company", "Data Type", "Year", "Field", "Valid", "Confidence", "Severity", "Message", "Suggestion"]
        ws.append(headers)

        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Data
        for validation in report_data["detailed_validations"]:
            company = validation["company"]
            result = validation["validation"]

            ws.append([
                company,
                validation.get("data_type", ""),
                validation.get("year", ""),
                result.get("field_name", ""),
                "Yes" if result.get("is_valid") else "No",
                f"{result.get('confidence_score', 0):.1%}",
                result.get("severity", ""),
                result.get("message", ""),
                result.get("suggestion", "")
            ])

    async def close(self):
        """Close resources."""
        await self.source_verifier.close()
