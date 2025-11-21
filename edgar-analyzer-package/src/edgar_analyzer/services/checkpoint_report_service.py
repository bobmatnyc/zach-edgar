"""Report generation service that works with checkpoint data."""

import json
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
import structlog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from edgar_analyzer.config.settings import ConfigService
from edgar_analyzer.models.intermediate_data import AnalysisCheckpoint, CompanyExtractionData

logger = structlog.get_logger(__name__)


class CheckpointReportService:
    """Report generation service that works with checkpoint intermediate data."""

    def __init__(self, config: ConfigService):
        """Initialize checkpoint report service."""
        self._config = config
        self._output_dir = Path(config.settings.output_dir)
        self._output_dir.mkdir(parents=True, exist_ok=True)

        logger.info("Checkpoint report service initialized", output_dir=str(self._output_dir))

    async def generate_excel_report(
        self,
        checkpoint: AnalysisCheckpoint,
        output_filename: Optional[str] = None
    ) -> Path:
        """Generate Excel report from checkpoint data."""

        if output_filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"fortune500_analysis_{checkpoint.target_year}_{timestamp}.xlsx"

        output_path = self._output_dir / output_filename

        try:
            # Create Excel workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Create main analysis sheet
            await self._create_main_analysis_sheet(wb, checkpoint)

            # Create summary sheet
            await self._create_summary_sheet(wb, checkpoint)

            # Create company details sheet
            await self._create_company_details_sheet(wb, checkpoint)

            # Create extraction log sheet
            await self._create_extraction_log_sheet(wb, checkpoint)

            # Save workbook
            wb.save(output_path)

            logger.info(
                "Excel report generated",
                filepath=str(output_path),
                companies=len(checkpoint.companies),
                completed=checkpoint.completed_companies
            )

            return output_path

        except Exception as e:
            logger.error("Failed to generate Excel report", error=str(e))
            raise

    def _calculate_summary_statistics(self, checkpoint: AnalysisCheckpoint) -> Dict:
        """Calculate summary statistics from checkpoint data."""
        completed_companies = checkpoint.get_completed_companies()

        if not completed_companies:
            return {
                "total_companies": checkpoint.total_companies,
                "companies_analyzed": 0,
                "companies_with_higher_compensation": 0,
                "percentage_higher_compensation": 0.0,
                "average_compensation_tax_ratio": 0.0,
                "target_year": checkpoint.target_year
            }

        # Calculate companies where compensation exceeds tax
        companies_with_higher_comp = 0
        total_ratios = []

        for company in completed_companies:
            target_year_ratio = company.compensation_vs_tax_ratios.get(checkpoint.target_year)
            if target_year_ratio and target_year_ratio > 1.0:
                companies_with_higher_comp += 1

            if target_year_ratio:
                total_ratios.append(target_year_ratio)

        return {
            "total_companies": checkpoint.total_companies,
            "companies_analyzed": len(completed_companies),
            "companies_with_higher_compensation": companies_with_higher_comp,
            "percentage_higher_compensation": (companies_with_higher_comp / len(completed_companies)) * 100,
            "average_compensation_tax_ratio": sum(total_ratios) / len(total_ratios) if total_ratios else 0.0,
            "target_year": checkpoint.target_year,
            "report_date": datetime.now()
        }

    def _convert_company_to_dict(self, company: CompanyExtractionData) -> Dict:
        """Convert company extraction data to dictionary format."""
        return {
            "company_info": {
                "cik": company.cik,
                "name": company.name,
                "ticker": company.ticker,
                "fortune_rank": company.fortune_rank,
                "sector": company.sector,
                "industry": company.industry
            },
            "extraction_metadata": {
                "status": company.status.value,
                "extraction_start_time": company.extraction_start_time.isoformat() if company.extraction_start_time else None,
                "extraction_end_time": company.extraction_end_time.isoformat() if company.extraction_end_time else None,
                "retry_count": company.retry_count
            },
            "financial_data": {
                "tax_data_by_year": company.tax_data,
                "compensation_data_by_year": company.compensation_data,
                "total_compensation_by_year": {
                    str(year): float(amount) for year, amount in company.total_compensation_by_year.items()
                },
                "compensation_vs_tax_ratios": {
                    str(year): ratio for year, ratio in company.compensation_vs_tax_ratios.items()
                }
            }
        }

    async def _create_main_analysis_sheet(self, workbook: Workbook, checkpoint: AnalysisCheckpoint) -> None:
        """Create main analysis sheet with company data."""
        ws = workbook.create_sheet("Analysis Results")

        # Create DataFrame from checkpoint data
        df = self._create_analysis_dataframe(checkpoint)

        # Add title
        title = f"Fortune 500 Executive Compensation vs Tax Analysis ({checkpoint.target_year})"
        ws.append([title])
        ws.merge_cells('A1:O1')

        # Style title
        title_cell = ws['A1']
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add empty row
        ws.append([])

        # Add data
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Style header row
        header_row = 3
        for cell in ws[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = chr(64 + column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_analysis_dataframe(self, checkpoint: AnalysisCheckpoint) -> pd.DataFrame:
        """Create analysis DataFrame from checkpoint data."""
        data = []
        completed_companies = checkpoint.get_completed_companies()

        for company in completed_companies:
            # Get data for target year
            target_year = checkpoint.target_year
            tax_data = company.tax_data.get(target_year, {})
            total_comp = float(company.total_compensation_by_year.get(target_year, 0))
            comp_tax_ratio = company.compensation_vs_tax_ratios.get(target_year)

            row = {
                'Rank': company.fortune_rank or 999,
                'Company': company.name,
                'Ticker': company.ticker or 'N/A',
                'Sector': company.sector or 'Unknown',
                'Industry': company.industry or 'Unknown',
                'Tax Expense (M)': tax_data.get('total_tax_expense', 0) / 1_000_000 if tax_data.get('total_tax_expense') else 0,
                'Executive Comp (M)': total_comp / 1_000_000,
                'Comp/Tax Ratio': comp_tax_ratio,
                'Comp > Tax': 'Yes' if comp_tax_ratio and comp_tax_ratio > 1 else 'No',
                'Effective Tax Rate': tax_data.get('effective_tax_rate'),
                'Extraction Status': company.status.value,
                'Data Years': len(company.tax_data)
            }

            # Add historical data columns
            for year in sorted(checkpoint.analysis_years):
                year_tax = company.tax_data.get(year, {}).get('total_tax_expense', 0)
                year_comp = float(company.total_compensation_by_year.get(year, 0))

                row[f'Tax {year} (M)'] = year_tax / 1_000_000 if year_tax else 0
                row[f'Comp {year} (M)'] = year_comp / 1_000_000

            data.append(row)

        df = pd.DataFrame(data)

        # Sort by Fortune ranking
        if not df.empty:
            df = df.sort_values('Rank', ascending=True)

        return df

    async def _create_summary_sheet(self, workbook: Workbook, checkpoint: AnalysisCheckpoint) -> None:
        """Create summary sheet with key statistics."""
        ws = workbook.create_sheet("Summary")

        stats = self._calculate_summary_statistics(checkpoint)

        summary_data = [
            ["Fortune 500 Analysis Summary", ""],
            ["Analysis ID", checkpoint.analysis_id],
            ["Target Year", checkpoint.target_year],
            ["Analysis Period", f"{min(checkpoint.analysis_years)}-{max(checkpoint.analysis_years)}"],
            ["Total Companies", checkpoint.total_companies],
            ["Successfully Analyzed", stats["companies_analyzed"]],
            ["Failed Extractions", checkpoint.failed_companies],
            ["Success Rate", f"{checkpoint.success_rate:.1f}%"],
            ["Companies with Higher Compensation", stats["companies_with_higher_compensation"]],
            ["Percentage with Higher Compensation", f"{stats['percentage_higher_compensation']:.1f}%"],
            ["Average Comp/Tax Ratio", f"{stats['average_compensation_tax_ratio']:.2f}"],
            ["Created", checkpoint.created_at.strftime("%Y-%m-%d %H:%M:%S")],
            ["Last Updated", checkpoint.last_updated.strftime("%Y-%m-%d %H:%M:%S")],
        ]

        for row in summary_data:
            ws.append(row)

        # Style header
        header_font = Font(bold=True, size=14)
        ws['A1'].font = header_font

    async def _create_company_details_sheet(self, workbook: Workbook, checkpoint: AnalysisCheckpoint) -> None:
        """Create detailed company data sheet."""
        ws = workbook.create_sheet("Company Details")

        # Headers
        headers = ["CIK", "Company", "Status", "Start Time", "End Time", "Retries", "Error Message"]
        ws.append(headers)

        # Data
        for company in checkpoint.companies:
            row = [
                company.cik,
                company.name,
                company.status.value,
                company.extraction_start_time.strftime("%Y-%m-%d %H:%M:%S") if company.extraction_start_time else "",
                company.extraction_end_time.strftime("%Y-%m-%d %H:%M:%S") if company.extraction_end_time else "",
                company.retry_count,
                company.error_message or ""
            ]
            ws.append(row)

        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)

    async def _create_extraction_log_sheet(self, workbook: Workbook, checkpoint: AnalysisCheckpoint) -> None:
        """Create extraction log sheet."""
        ws = workbook.create_sheet("Extraction Log")

        log_data = [
            ["Extraction Log", ""],
            ["Analysis ID", checkpoint.analysis_id],
            ["Created", checkpoint.created_at.isoformat()],
            ["Last Updated", checkpoint.last_updated.isoformat()],
            ["Configuration", ""],
        ]

        # Add config data
        for key, value in checkpoint.config.items():
            log_data.append([f"  {key}", str(value)])

        log_data.append(["", ""])
        log_data.append(["Global Errors", ""])

        # Add global errors
        for i, error in enumerate(checkpoint.global_errors):
            log_data.append([f"Error {i+1}", str(error)])

        for row in log_data:
            ws.append(row)

        # Style header
        header_font = Font(bold=True, size=14)
        ws['A1'].font = header_font

    async def generate_json_report(
        self,
        checkpoint: AnalysisCheckpoint,
        output_filename: Optional[str] = None
    ) -> Path:
        """Generate JSON report from checkpoint data."""

        if output_filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"fortune500_analysis_{checkpoint.target_year}_{timestamp}.json"

        output_path = self._output_dir / output_filename

        try:
            # Create comprehensive report data
            report_data = {
                "metadata": {
                    "analysis_id": checkpoint.analysis_id,
                    "target_year": checkpoint.target_year,
                    "analysis_years": checkpoint.analysis_years,
                    "generated_at": datetime.now().isoformat(),
                    "total_companies": checkpoint.total_companies,
                    "completed_companies": checkpoint.completed_companies,
                    "failed_companies": checkpoint.failed_companies,
                    "success_rate": checkpoint.success_rate,
                    "progress_percentage": checkpoint.progress_percentage
                },
                "summary_statistics": self._calculate_summary_statistics(checkpoint),
                "companies": [
                    self._convert_company_to_dict(company)
                    for company in checkpoint.get_completed_companies()
                ],
                "failed_companies": [
                    {
                        "cik": company.cik,
                        "name": company.name,
                        "error_message": company.error_message,
                        "retry_count": company.retry_count
                    }
                    for company in checkpoint.get_failed_companies()
                ],
                "extraction_log": {
                    "created_at": checkpoint.created_at.isoformat(),
                    "last_updated": checkpoint.last_updated.isoformat(),
                    "config": checkpoint.config,
                    "global_errors": checkpoint.global_errors
                }
            }

            # Save to JSON file
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(report_data, f, indent=2, default=str)

            logger.info(
                "JSON report generated",
                filepath=str(output_path),
                companies=len(checkpoint.companies),
                completed=checkpoint.completed_companies
            )

            return output_path

        except Exception as e:
            logger.error("Failed to generate JSON report", error=str(e))
            raise