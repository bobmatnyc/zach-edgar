"""Report generation service implementation."""

import json
from datetime import datetime
from pathlib import Path
from typing import List

import pandas as pd
import structlog
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from edgar_analyzer.config.settings import ConfigService
from edgar_analyzer.models.company import AnalysisReport, CompanyAnalysis
from edgar_analyzer.services.interfaces import IDataExtractionService, IReportService

logger = structlog.get_logger(__name__)


class ReportService(IReportService):
    """Report generation service implementation."""

    def __init__(
        self,
        data_extraction_service: IDataExtractionService,
        config: ConfigService
    ):
        """Initialize report service."""
        self._data_extraction = data_extraction_service
        self._config = config
        self._output_dir = Path(config.settings.output_dir)
        self._output_dir.mkdir(parents=True, exist_ok=True)

        logger.info("Report service initialized", output_dir=str(self._output_dir))

    async def generate_analysis_report(
        self, companies: List[str], year: int
    ) -> AnalysisReport:
        """Generate analysis report for multiple companies."""
        logger.info("Starting analysis report generation", companies=len(companies), year=year)

        report = AnalysisReport(target_year=year)

        for cik in companies:
            try:
                logger.info("Analyzing company", cik=cik)

                # Extract company analysis
                company_analysis = await self._data_extraction.extract_company_analysis(cik, year)

                if company_analysis:
                    report.add_company_analysis(company_analysis)
                    logger.info(
                        "Company analysis added to report",
                        cik=cik,
                        company=company_analysis.company.name
                    )
                else:
                    logger.warning("Failed to analyze company", cik=cik)

            except Exception as e:
                logger.error("Error analyzing company", cik=cik, error=str(e))
                continue

        logger.info(
            "Analysis report generation completed",
            total_companies=report.total_companies,
            year=year
        )

        return report

    async def export_to_excel(self, report: AnalysisReport, filepath: str) -> None:
        """Export report to Excel file."""
        try:
            # Create DataFrame from report data
            df = self._create_report_dataframe(report)

            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"Executive Compensation vs Tax Analysis {report.target_year}"

            # Add header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            # Write data to worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            # Style header row
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Add summary sheet
            self._add_summary_sheet(wb, report)

            # Save workbook
            output_path = self._output_dir / filepath
            wb.save(output_path)

            logger.info("Excel report exported", filepath=str(output_path))

        except Exception as e:
            logger.error("Failed to export Excel report", filepath=filepath, error=str(e))
            raise

    async def export_to_json(self, report: AnalysisReport, filepath: str) -> None:
        """Export report to JSON file."""
        try:
            output_path = self._output_dir / filepath

            # Convert report to dict
            report_data = report.dict()

            # Write to JSON file
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(report_data, f, indent=2, default=str)

            logger.info("JSON report exported", filepath=str(output_path))

        except Exception as e:
            logger.error("Failed to export JSON report", filepath=filepath, error=str(e))
            raise

    def _create_report_dataframe(self, report: AnalysisReport) -> pd.DataFrame:
        """Create DataFrame from analysis report."""
        data = []

        for company_analysis in report.companies:
            company = company_analysis.company

            # Get latest tax expense
            tax_expense = company_analysis.latest_tax_expense
            tax_amount = tax_expense.total_tax_expense if tax_expense else 0

            # Get total executive compensation for target year
            total_comp_by_year = company_analysis.total_executive_compensation
            exec_compensation = total_comp_by_year.get(report.target_year, 0)

            # Calculate ratio
            ratio = None
            if tax_amount > 0:
                ratio = float(exec_compensation / tax_amount)

            # Determine if compensation exceeds tax
            exceeds_tax = exec_compensation > tax_amount if tax_amount > 0 else False

            row = {
                'Company Name': company.name,
                'Ticker': company.ticker or 'N/A',
                'CIK': company.cik,
                'Industry': company.industry or 'N/A',
                'Sector': company.sector or 'N/A',
                'Fortune Rank': company.fortune_rank or 'N/A',
                f'Executive Compensation {report.target_year}': float(exec_compensation),
                f'Tax Expense {report.target_year}': float(tax_amount),
                'Compensation/Tax Ratio': ratio,
                'Compensation Exceeds Tax': exceeds_tax,
                'Number of Executives': len(company_analysis.executive_compensations)
            }

            data.append(row)

        df = pd.DataFrame(data)

        # Sort by compensation amount (descending)
        if not df.empty:
            comp_col = f'Executive Compensation {report.target_year}'
            df = df.sort_values(comp_col, ascending=False)

        return df

    def _add_summary_sheet(self, workbook: Workbook, report: AnalysisReport) -> None:
        """Add summary sheet to workbook."""
        ws = workbook.create_sheet("Summary")

        # Add summary statistics
        stats = report.summary_statistics

        summary_data = [
            ["Analysis Summary", ""],
            ["Report Date", stats["report_date"].strftime("%Y-%m-%d %H:%M:%S")],
            ["Target Year", stats["target_year"]],
            ["Total Companies Analyzed", stats["total_companies"]],
            ["Companies with Higher Compensation", stats["companies_with_higher_compensation"]],
            ["Percentage with Higher Compensation", f"{stats['percentage_higher_compensation']:.1f}%"],
        ]

        for row in summary_data:
            ws.append(row)

        # Style summary sheet
        header_font = Font(bold=True, size=14)
        ws['A1'].font = header_font

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width